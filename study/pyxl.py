import sys
from PyQt5.QtWidgets import *
import openpyxl
from openpyxl import *

class Openfile(QWidget):
    
    def __init__(self):
        super().__init__()
        self.initUI()
        self.read()
        
    def initUI(self):
        
        self.texta = QTextBrowser(self)
        self.textb = QTextBrowser(self)
        self.textc = QTextBrowser(self)
        self.textd = QTextBrowser(self)
        
        self.vbox = QVBoxLayout()
        self.vbox.addWidget(self.texta)
        self.vbox.addWidget(self.textb)
        self.vbox.addWidget(self.textc)
        self.vbox.addWidget(self.textd)
        
        self.setLayout(self.vbox)
        
        self.setGeometry(100, 150, 950, 800)
        self.show()
        
    def read(self):
        wb= openpyxl.load_workbook('./scenario_1.xlsx')
        List = [wb.sheetnames[0],
                     wb.sheetnames[1],wb.sheetnames[2]]
        
        ws1 = wb['CLU_01_20ms']
        ws2 = wb['CLU_02_100ms']
        ws3 = wb['WHL_01_10ms'] #시트정하기
        
        names = [] 
        
        for i in range(2,503):
            timelist = ws1.cell(row=i, column=1).value # 열정하기
            names.append(timelist)
        
        # print(names[25])
        
        for cell_obj in list(ws1.rows)[0]:
            signallist1= cell_obj.value # 특정행 불러오기 - sinallist
            # print(signallist1) # 행 6개
            
        for cell_obj in list(ws2.rows)[0]:
            signallist2= cell_obj.value # 특정행 불러오기 - sinallist
            # print(signallist2) # 행 4개
            
        for cell_obj in list(ws3.rows)[0]:
            signallist3= cell_obj.value # 특정행 불러오기 - sinallist
            # print(signallist3) # 행 11개
            
        Signallist = [signallist1, signallist2, signallist3]
        self.texta.append(str(List))
        self.textb.append(str(Signallist))
        self.textc.append(str(signallist2))
        self.textd.append(str(signallist3))
        
if __name__ == '__main__':
    app = QApplication(sys.argv)
    myapp = Openfile()
    myapp.show()
    app.exec_()
        
import openpyxl
from openpyxl import *
class Signal():
    wb= openpyxl.load_workbook('./scenario_1.xlsx')
    ws1 = wb['CLU_01_20ms']
    ws2 = wb['CLU_02_100ms']
    ws3 = wb['WHL_01_10ms']
    
    name_col = ws1['A']
    names = []

    for j in range(2,7):
        for i in range(2,503):
            timelist1 = ws1.cell(row=i, column=j).value 
            names.append(timelist1)
            signal = names[0:501]
            signal1 = names[501:1002]
            signal2 = names[1002:1503]
            signal3 = names[1503:2004]
    print(signal3)
    
    
    # for i in range(1,503):
    #     timelist2 = ws2.cell(row=i, column=6).value 
    #     names.append(timelist2)
    
    # for i in range(1,503):
    #     timelist3 = ws3.cell(row=i, column=6).value 
    #     names.append(timelist3)
        # print(timelist3)
        
    # for cell in name_col:
    #     names.append(cell.value)
    # print(names)